import os
from tkinter import *
from tkinter import ttk
from openpyxl import Workbook, load_workbook
import tkinter.filedialog as ttfiledialog
import re
import json
import pyperclip as MyClipboard
import pyautogui
import time
import requests
import chardet
import shutil

def gooletranslate(key, source_language='CN-简体', target_language='EN-英语'):
    target_language_dict={
    "CN-简体":"ZhCN",
    "EN-英语":"En",
    "TW-繁体":"ZhTw",
    "ES-西班牙":"Es",
    "PL-波兰":"Pl",
    "PT-葡萄牙":"Pt",
    "RU-俄语":"Ru",
    "FR-法语":"Fr",
    "DE-德语":"De",
    "IT-意大利":"It",
    "JA-日语":"Ja",
    "FI-芬兰":"Fi",
    "VN-越南":"Vi",
    "KR-韩语":"Ko",
    "AR-阿拉伯":"Ar",
    "TR-土耳其":"Tr",
    "TH-泰语":"Th",
    "HU-匈牙利":"Hu",
    "EL-希腊":"El",
    "NL-荷兰":"Nl",
    "NO-挪威":"No",
    "MA-马来西亚":"My",
    "FA-波斯":"Fa",
    "DA-丹麦":"Da",
    "RO-罗马尼亚":"Ro",
    "BR-葡萄牙(巴西)":"Br",
    "ID-印尼":"Id"
    }
    source_language_dict={
    "CN-简体":"ZhCN",
    "EN-英语":"En",
    "TW-繁体":"ZhTw",
    "ES-西班牙":"Es",
    "PL-波兰":"Pl",
    "PT-葡萄牙":"Pt",
    "RU-俄语":"Ru",
    "FR-法语":"Fr",
    "DE-德语":"De",
    "IT-意大利":"It",
    "JA-日语":"Ja",
    "FI-芬兰":"Fi",
    "VN-越南":"Vi",
    "KR-韩语":"Ko",
    "AR-阿拉伯":"Ar",
    "TR-土耳其":"Tr",
    "TH-泰语":"Th",
    "HU-匈牙利":"Hu",
    "EL-希腊":"El",
    "NL-荷兰":"Nl",
    "NO-挪威":"No",
    "MA-马来西亚":"My",
    "FA-波斯":"Fa",
    "DA-丹麦":"Da",
    "RO-罗马尼亚":"Ro",
    "BR-葡萄牙(巴西)":"Br",
    "ID-印尼":"Id"
    }
    if target_language not in target_language_dict:
        return "不支持该种语言翻译"
    target_language = target_language_dict[target_language]
    translate_url = "http://19.87.8.22:36604/wordsTranslationDict/translate"
    translate_userinfo = "eyJhY2Nlc3NfdG9rZW4iOiI1MjIzZjZiNTgxMmY0NmMyYTJhMjgyM2Q4MGJhNWI1NiIsInJlZnJlc2hfdG9rZW4iOiI4NWM4MDcwOWI2NzA0YTk5OWJmYzE5ZGRiODM5MWE5NyIsImRvbWFpbiI6Ilh0b29sdGVjaCIsImlkIjoiMjMxMDE2MzQiLCJuaWNrX25hbWUiOiLogpblmInkvJ8iLCJhdmF0YXJfdXJsIjpudWxsLCJwaG9uZSI6bnVsbCwiaXNCaW5kRGluZ3RhbGsiOmZhbHNlLCJleHBpcmUiOjE3MjA1MjI4NzMyNDEsInJlcyI6W119"

    headers = {'userinfo': translate_userinfo,
               'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                             'Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0'}
    payload = {
        'CC': 'PC',
        'SourceLanguage': "ZhCN",
        'TargetLanguage': target_language,
        'words[0]': key  
    }
    response = requests.post(translate_url, headers=headers, data=payload)
    message = ""
    if response.status_code == 200:
        response_data = response.json()
        try:
            translation_result = response_data['data']['WordsResults'][0]['Result']
            message = translation_result
        except Exception as e:
            message = "--Undefined--" + str(e)
    return message
    
class myTranslator:
    def __init__(self):
        self.NoTranFlah = "--Undefined--"
        self.NOTranDataDict = {}
        self.logInfo=[]
        self.UNIT_dict = {}
        self.chinese_to_english = self.load_chinese_to_english_mapping()
        self.encodingDict={
        "HU-匈牙利":"1250",
        "RO-罗马尼亚":"1250",
        "CN-简体":"936",
        "TW-繁体":"950",
        "CS":"1250",
        "DA-丹麦":"1252",
        "DE-德语":"1252",
        "EL-希腊":"1253",
        "EN-英语":"utf-8",
        "ES-西班牙":"1252",
        "AR-阿拉伯":"1256",
        "FI-芬兰":"1252",
        "FR-法语":"1252",
        "IT-意大利":"1252",
        "JA-日语":"932",
        "KR-韩语":"949",
        "NL-荷兰":"1252",
        "NO-挪威":"1142(ISO-8859-10)",
        "PL-波兰":"1250",
        "PT-葡萄牙":"1252",
        "RU-俄语":"1251",
        "SV":"1252",
        "TR-土耳其":"1254",
        "FA-波斯":"1256",
        "MA-马来西亚":"utf-8",#
        "SR":"1252",
        "VN-越南":"1258",
        "ID-印尼":"Unicode",
        "BR-葡萄牙(巴西)":"1252"}
    
    def on_closing(self):
        self.MyGUI.after_cancel(self.updateId)
        self.MyGUI.destroy()

    def AddValue2Dict(self,k,v):
        if k not in self.ExistDataDict:
            if len(v)!=0:
                self.ExistDataDict[k] = v

    def creatFile(self):
        self.strVehicelxls = self.DataPath + "翻译词典.xlsx"
        folder = os.path.exists(self.strVehicelxls)
        if not folder:
            xls = Workbook()
            sht1 = xls.active
            sht1.title = "已经校验翻译"
            sht2 = xls.create_sheet("现有翻译")
            sht3 = xls.create_sheet("有道翻译")
            sht4 = xls.create_sheet("未翻译")
            
            # 设置字体格式
            from openpyxl.styles import Font
            Font0 = Font(name="Times New Roman", color="FF0000", bold=True)
            for sheet in [sht1, sht2, sht3, sht4]:
                for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.font = Font0

            listLangue = ["CN-简体","EN-英语","TW-繁体","ES-西班牙","PL-波兰","PT-葡萄牙","RU-俄语","FR-法语","DE-德语","IT-意大利","JA-日语","FI-芬兰","VN-越南","KR-韩语","AR-阿拉伯","TR-土耳其","TH-泰语","HU-匈牙利","EL-希腊","NL-荷兰","NO-挪威","MA-马来西亚","FA-波斯","DA-丹麦","RO-罗马尼亚","BR-葡萄牙(巴西)","ID-印尼"]
            lit = 0
            for langue in listLangue:
                sht1.cell(row=1, column=lit+1, value=langue)
                lit=lit+1

            lit = 0
            for langue in listLangue:
                sht2.cell(row=1, column=lit+1, value=langue)
                lit=lit+1
                
            lit = 0
            for langue in listLangue:
                sht3.cell(row=1, column=lit+1, value=langue)
                lit=lit+1
                
            sht4.cell(row=1, column=1, value="")
            xls.save(self.strVehicelxls)
        
        if self.CheckAll.get() != 1:
            folder = os.path.exists(self.strOutPath)
            if not folder:
                os.makedirs(self.strOutPath)
            
        folder = os.path.exists("UNIT.json")
        if not folder:
            with open("UNIT.json","w") as CreateJS:
                json.dump(self.UNIT_dict,CreateJS,ensure_ascii=False)
                
        with open("UNIT.json",'r',encoding = "utf-8")as json_Read:
            self.UNIT_dict = json.load(json_Read)

    def xls2dict(self):
        self.UpdataLog("*************************************正在翻译语言："+self.TargeLanguage+"*************************************")
        self.CheckedDataDict = {}
        self.ExistDataDict = {}
        self.YodaoDataDict = {}
        mywork = load_workbook(self.strVehicelxls)
        sheet = mywork["已经校验翻译"]
        rows = sheet.max_row
        cols = sheet.max_column
        for col in range(2, cols + 1):
            language = sheet.cell(row=1, column=col).value
            if language.find(self.TargeLanguage) != -1:
                cols = col
                break
        for row in range(2, rows + 1):
            cnData = str(sheet.cell(row=row, column=1).value)
            targData = str(sheet.cell(row=row, column=cols).value).replace("\n", "\\n")
            self.CheckedDataDict[cnData] = targData
    
        sheet = mywork["现有翻译"]
        rows = sheet.max_row
        cols = sheet.max_column
        for col in range(2, cols + 1):
            language = sheet.cell(row=1, column=col).value
            if language.find(self.TargeLanguage) != -1:
                cols = col
                break
        for row in range(2, rows + 1):
            cnData = str(sheet.cell(row=row, column=1).value)
            targData = str(sheet.cell(row=row, column=cols).value)
            self.ExistDataDict[cnData] = targData

        sheet = mywork["有道翻译"]
        rows = sheet.max_row
        cols = sheet.max_column
        for col in range(2, cols + 1):
            language = sheet.cell(row=1, column=col).value
            if language.find(self.TargeLanguage) != -1:
                cols = col
                break
        for row in range(2, rows + 1):
            cnData = str(sheet.cell(row=row, column=1).value)
            targData = str(sheet.cell(row=row, column=cols).value)
            if len(targData) != 0:
                self.YodaoDataDict[cnData] = targData

    def creatUI(self):
        self.bTrans = 1
        self.f_path=""
        self.MyGUI = Tk()
        self.MyGUI.title("翻译文档")
        self.MyGUI.geometry('947x482+10+10')
        #语言选择
        LangueTips = Label(self.MyGUI,justify = 'left',anchor='n', text='选择翻译的语言：')
        LangueTips.grid(row=0, column=0)
        self.LangueSelect = ttk.Combobox(self.MyGUI,width=15)
        self.LangueSelect['value'] = ("国产五种-英西俄法阿","ALL-所有","EN-英语","TW-繁体","ES-西班牙","PL-波兰","PT-葡萄牙","RU-俄语","FR-法语","DE-德语","IT-意大利","JA-日语","FI-芬兰","VN-越南","KR-韩语","AR-阿拉伯","TR-土耳其","TH-泰语","HU-匈牙利","EL-希腊","NL-荷兰","NO-挪威","MA-马来西亚","FA-波斯","DA-丹麦","RO-罗马尼亚","BR-葡萄牙(巴西)","ID-印尼")
        self.LangueSelect.current(0)
        self.LangueSelect.grid(row=1, column=0, sticky='NS')
        
        #对象选择
        ObjTips = Label(self.MyGUI,justify = 'left', text='选择翻译的对象：')
        ObjTips.grid(row=0, column=1)
        self.ObjSelect = ttk.Combobox(self.MyGUI,width=15)
        self.ObjSelect['value'] = ("所有.txt","ALL","DTC.txt","TEXT.txt","DS.txt","DTC_H.txt","MENU.txt","ROOT.txt","QuickInfo.txt","EXCEL.xls")
        self.ObjSelect.current(0)
        self.ObjSelect.grid(row=1, column=1, sticky='NS')
        
        #选择文件夹
        FileTips = Label(self.MyGUI,justify = 'left', text='选择翻译的目录：')
        FileTips.grid(row=0, column=2)
        self.DataSelect = ttk.Combobox(self.MyGUI,width=45)
        self.DataSelect['value']=("选择文件夹")
        self.DataSelect.grid(row=1, column=2, sticky='NS')
        
        youdao = Label(self.MyGUI,justify = 'left', text="使用翻译平台：")
        youdao.grid(row=0, column=3)
        self.Selectyoudao = ttk.Combobox(self.MyGUI,width=10)
        self.Selectyoudao['value']=("是","否")
        self.Selectyoudao.current(1)
        self.Selectyoudao.grid(row=1, column=3, sticky='NS')
        
        FileSelect = Button(self.MyGUI,text="开始翻译",bg='lightblue',command = self.TranStart)
        FileSelect.grid(row=1,column=4, sticky='EW')
        
        #FileSelect = Button(self.MyGUI,text="检查乱码",bg='lightblue',command = self.CheckGarbledCode)
        #FileSelect.grid(row=1,column=5, sticky='EW')

        self.FormatAdjust = Button(self.MyGUI,text="格式调整",bg='lightblue',command = self.CheckFormatStart)
        self.FormatAdjust.grid(row=1,column=5, sticky='EW')

        FileSelect = Button(self.MyGUI,text="提取翻译内容",bg='lightblue',command = self.GetTransInfo)
        FileSelect.grid(row=1,column=6, sticky='EW')
        
        self.progress = ttk.Progressbar(self.MyGUI, orient="horizontal", length=400, mode="determinate")
        self.progress.grid(row=3, column=0, columnspan=5, pady=5)  # 添加进度条到界面

        FileSelect = Button(self.MyGUI,text="提取ROOT工程IDM",bg='lightblue',command = self.GetIDM)
        FileSelect.grid(row=3,column=4, sticky='EW')

        #添加一个开关，用于获取文件夹内所有文件
        self.CheckAll = IntVar()
        self.CheckAll.trace_add("write", self.toggle_controls)
        CheckAll = Checkbutton(self.MyGUI, text="迭代读取所有文件", variable=self.CheckAll, onvalue=1, offvalue=0)
        CheckAll.grid(row=3, column=5, columnspan=2, sticky='EW')

        LogTips = Label(self.MyGUI,justify = 'left',text = "实时日志：")
        LogTips.grid(row=3,column = 0)
        self.LogBox = Text(self.MyGUI, width=133, height=30)
        self.LogBox.place(x=5,y=85)

        self.MyGUI.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.UpDataGUI()

    def toggle_controls(self, *args):
        # 根据CheckAll的值来启用或禁用控件
        if self.CheckAll.get() == 1:
            self.ObjSelect.config(state='disabled')
            self.FormatAdjust.config(state='disabled')
        else:
            self.ObjSelect.config(state='normal')
            self.FormatAdjust.config(state='normal')

    def UpDataGUI(self):
        self.MyGUI.update()
        self.updateId = self.MyGUI.after(1000,self.UpDataGUI)
        if self.Selectyoudao.get()=="是":
            self.Toyoudao = True
        else:
            self.Toyoudao = False
        if self.DataSelect.get()=="选择文件夹":
            self.MyGUI.after_cancel( self.updateId)
            self.GetPath()

    def GetPath(self):
        self.DataPath = ttfiledialog.askdirectory()
        self.DataPath = self.DataPath+"/"
        ipos1 = self.DataPath.find(r"GS_")
        ipos2 = self.DataPath.find(r"/data")
        self.strVehicel = self.DataPath[ipos1+3:ipos2]
        self.DataSelect.set(self.DataPath)
        self.UpDataGUI()
    
    def TranStart(self):
        if self.CheckAll.get() == 1:
            self.TranStartAll()
        else:
            if self.strVehicel=="":
                return 0
            self.LangueSelect.update()
            LangueList = []
            if self.LangueSelect.get()=="ALL-所有":
                LangueList = ["EN-英语","TW-繁体","ES-西班牙","PL-波兰","PT-葡萄牙","RU-俄语","FR-法语","DE-德语","IT-意大利","JA-日语","FI-芬兰","VN-越南","KR-韩语","AR-阿拉伯","TR-土耳其","TH-泰语","HU-匈牙利","EL-希腊","NL-荷兰","NO-挪威","MA-马来西亚","FA-波斯","DA-丹麦","RO-罗马尼亚","BR-葡萄牙(巴西)","ID-印尼"]
            elif self.LangueSelect.get()=="国产五种-英西俄法阿":
                LangueList = ["EN-英语","ES-西班牙","RU-俄语","FR-法语","AR-阿拉伯"]
            else:
                LangueList = [self.LangueSelect.get()]
            
            self.ObjSelect.update()
            FileTypeList = []
            if self.ObjSelect.get()=="ALL":
                FileTypeList = ["DTC.txt","TEXT.txt","DS.txt","DTC_H.txt","MENU.txt","ROOT.txt","EXCEL.xls","QuickInfo.txt"]
            elif self.ObjSelect.get()=="所有.txt":
                FileTypeList = ["ROOT.txt","MENU.txt","TEXT.txt","DS.txt","DTC_H.txt","DTC.txt","QuickInfo.txt"]
            else:
                FileTypeList = [self.ObjSelect.get()]
            
            for Langue in LangueList:
                self.strOutPath = self.DataPath+Langue+"/"
                self.TargeLanguage = Langue
                self.creatFile()
                self.xls2dict()
                for FileType in FileTypeList:
                    self.FileType = FileType
                    self.SelectFun()
            self.UpdataLog("*************************************END OF ALL*************************************"+"\n")
            self.mylog()

    def TranStartAll(self):
        #获取文件夹内的所有文件,如果还有子文件夹，递归
        file_list = []
        file_listCopy = []
        try:#没有选择文件夹
            for root, dirs, files in os.walk(self.DataPath):
                for file in files:#.txt文件
                    if (file.endswith(".txt") and file.find("_NeedTrans.txt") == -1) or file.endswith(".ini"):
                        file_list.append(os.path.join(root, file))
                    else:
                        file_listCopy.append(os.path.join(root, file))
        except:
            self.UpdataLog("********************************未选择文件夹********************************")
            return 0

        if self.strVehicel=="":
                return 0

        #去除DataPath末尾的“/”
        if self.DataPath[-1] == "/":
            self.Path = self.DataPath[:-1]

        self.LangueSelect.update()
        LangueList = []
        if self.LangueSelect.get()=="ALL-所有":
            LangueList = ["EN-英语","TW-繁体","ES-西班牙","PL-波兰","PT-葡萄牙","RU-俄语","FR-法语","DE-德语","IT-意大利","JA-日语","FI-芬兰","VN-越南","KR-韩语","AR-阿拉伯","TR-土耳其","TH-泰语","HU-匈牙利","EL-希腊","NL-荷兰","NO-挪威","MA-马来西亚","FA-波斯","DA-丹麦","RO-罗马尼亚","BR-葡萄牙(巴西)","ID-印尼"]
        elif self.LangueSelect.get()=="国产五种-英西俄法阿":
            LangueList = ["EN-英语","ES-西班牙","RU-俄语","FR-法语","AR-阿拉伯"]
        else:
            LangueList = [self.LangueSelect.get()]

        for Langue in LangueList:
            language_code = Langue.split('-')[0] if '-' in Langue else Langue
            target_folder = os.path.join(self.DataPath, f"{self.Path}_{language_code}")
            if not os.path.exists(target_folder):
                os.makedirs(target_folder)
                self.UpdataLog(f"创建目标文件夹: {target_folder}")
            self.TargeLanguage = Langue
            self.creatFile()
            self.xls2dict()
            total_files = len(file_list)
            self.progress["maximum"] = total_files
            self.progress["value"] = 0
            self.MyGUI.update()
            for file in file_list:
                file_target = os.path.join(target_folder, file.split("/")[-1])
                # 确保目标文件夹存在，如果不存在则创建
                file_target_dir = os.path.dirname(file_target)
                if not os.path.exists(file_target_dir):
                    os.makedirs(file_target_dir)
                if file.endswith(".ini"):
                    try:
                        with open(file, r"r", encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Read:
                            with open(file_target, r"w", encoding = self.encodingDict[self.TargeLanguage], errors='replace') as EN_Write:
                                for line in CN_Read:
                                    EN_Value = ""
                                    #去掉前后空格换行符
                                    line = line.strip()
                                    if contains_chinese(line):
                                        CN_Values = line.split("\t")
                                        EN_Value = self.getTargeValue(CN_Values[1])
                                        line = line.replace(CN_Values[1], EN_Value)
                                    line = line + "\n"
                                    line = self.replace_chinese_symbol(line)
                                    EN_Write.write(line)
                        self.progress["value"] += 1
                        self.MyGUI.update()
                    except Exception as e:
                        self.UpdataLog(f"翻译失败：{file} {e}") 
                else:
                    try:
                        with open(file, r"r", encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Read:
                            with open(file_target, r"w", encoding = self.encodingDict[self.TargeLanguage], errors='replace') as EN_Write:
                                for line in CN_Read:
                                    line = self.replace_chinese_symbol(line)
                                    #遇到line中存在乱码，则直接写入整行
                                    if has_invalid_characters(line, self.encodingDict["CN-简体"]):
                                        EN_Write.write(line)
                                        continue
                                    EN_Value = ""
                                    if contains_chinese(line):
                                        CN_Values = extract_chinese_from_ternary(line)
                                        for CN_Value in CN_Values:
                                            EN_Value = self.getTargeValue(CN_Value)
                                            if EN_Value == "":
                                                self.UpdataLog(f"翻译失败：{CN_Value}")
                                            else:
                                                line = replace_exact_chinese(line, CN_Value, EN_Value)
                                        EN_Write.write(line)  
                                    else:
                                        EN_Write.write(line)
                        self.progress["value"] += 1
                        self.MyGUI.update()
                    except Exception as e:
                        self.UpdataLog(f"翻译失败：{file} {e}") 
            for file in file_listCopy:
                if file.endswith(".xls") or file.endswith(".xlsx"):
                    continue
                file_target = os.path.join(target_folder, file.split("/")[-1])
                # 确保目标文件夹存在，如果不存在则创建
                file_target_dir = os.path.dirname(file_target)
                if not os.path.exists(file_target_dir):
                    os.makedirs(file_target_dir)
                try:
                    shutil.copy(file, file_target)
                except Exception as e:
                    self.UpdataLog(f"复制失败：{file} {e}") 
            self.UpdataLog(f"翻译完成：{target_folder}")
        self.UpdataLog("*************************************END OF ALL*************************************"+"\n")
        self.mylog()

    def create_target_folders(self):
        # 获取选择的语言
        selected_language = self.LangueSelect.get()
        if selected_language == "ALL-所有":
            language_list = ["EN-英语", "TW-繁体", "ES-西班牙", "PL-波兰", "PT-葡萄牙", "RU-俄语", "FR-法语", "DE-德语", "IT-意大利", "JA-日语", "FI-芬兰", "VN-越南", "KR-韩语", "AR-阿拉伯", "TR-土耳其", "TH-泰语", "HU-匈牙利", "EL-希腊", "NL-荷兰", "NO-挪威", "MA-马来西亚", "FA-波斯", "DA-丹麦", "RO-罗马尼亚", "BR-葡萄牙(巴西)", "ID-印尼"]
        elif selected_language == "国产五种-英西俄法阿":
            language_list = ["EN-英语", "ES-西班牙", "RU-俄语", "FR-法语", "AR-阿拉伯"]
        else:
            language_list = [selected_language]

        #去除DataPath末尾的“/”
        if self.DataPath[-1] == "/":
            self.Path = self.DataPath[:-1]

        # 遍历每个语言，创建对应的目标文件夹
        for language in language_list:
            language_code = language.split('-')[0] if '-' in language else language
            target_folder = os.path.join(self.DataPath, f"{self.Path}_{language_code}")  # 使用 strVehicel 和语言代码生成目标文件夹名
            if not os.path.exists(target_folder):
                os.makedirs(target_folder)
                self.UpdataLog(f"创建目标文件夹: {target_folder}")

    def UpDict(self):
        with open("UNIT.json","w",encoding = "utf-8") as CreateJS:
            json.dump(self.UNIT_dict,CreateJS,ensure_ascii=False)

    def UpProgramXls(self):
        # 加载现有的xlsx文件
        wb = load_workbook(self.strVehicelxls)
        
        # 获取各个工作表
        # sheet1 = wb["已经校验翻译"]
        sheet2 = wb["现有翻译"]
        sheet3 = wb["有道翻译"]
        sheet4 = wb["未翻译"]
        
        # 现有翻译
        targeCol = 0
        for col in range(1, sheet2.max_column + 1):
            language = sheet2.cell(row=1, column=col).value
            if language.startswith(self.TargeLanguage):
                targeCol = col
                break
        row = 2  # 从第二行开始写入数据
        for k, v in self.ExistDataDict.items():
            sheet2.cell(row=row, column=1, value=k)
            sheet2.cell(row=row, column=targeCol, value=v)
            row += 1
        
        # 有道翻译
        targeCol = 0
        for col in range(1, sheet3.max_column + 1):
            language = sheet3.cell(row=1, column=col).value
            if language.startswith(self.TargeLanguage):
                targeCol = col
                break
        row = 2  # 从第二行开始写入数据
        for k, v in self.YodaoDataDict.items():
            sheet3.cell(row=row, column=1, value=k)
            sheet3.cell(row=row, column=targeCol, value=v)
            row += 1
        
        # 无翻译
        for row in range(1, sheet4.max_row + 1):
            sheet4.cell(row=row, column=1, value="")  # 清空
            sheet4.cell(row=row, column=2, value="")  # 清空
        row = 1  # 从第一行开始写入数据
        for k, v in self.NOTranDataDict.items():
            sheet4.cell(row=row, column=1, value=k)
            sheet4.cell(row=row, column=2, value=v)
            row += 1
        
        try:
            wb.save(self.strVehicelxls)
        except Exception as e:
            self.UpdataLog("大佬，先关闭这个文件哈：" + self.strVehicelxls)
            print(f"保存文件时出错: {e}")

    def SelectFun(self):
        self.UpdataLog("*******************************************************开始翻译："+self.TargeLanguage+"_"+self.FileType)
        self.CNFile = self.DataPath+"CN_"+self.FileType
        self.CNFileNew = self.DataPath+"CN_"+"NEW_"+self.FileType
        Language = self.TargeLanguage.split("-")[0]
        self.OutFile = self.strOutPath+Language+"_"+self.FileType
        self.TargeFile = self.DataPath+Language+"_"+self.FileType
        if self.FileType == "DTC.txt":
            self.CNDTCToOther()
        elif self.FileType == "DTC_H.txt":
            self.CNDTC_HToOther()
        elif self.FileType == "TEXT.txt":
            self.CNTextToOther()
        elif self.FileType == "DS.txt":
            self.CNDSToOther()
        elif self.FileType == "MENU.txt":
            self.CMenuToOther()
        elif self.FileType == "ROOT.txt":
            self.CRootToOther()
        elif self.FileType == "QuickInfo.txt":
            self.CQuickInfoToOther()
        self.UpDict()
        #self.UpProgramXls()
        self.UpdataLog("*******************************************************成功翻译："+Language+"_"+self.FileType+"\n")

    def Myfilter(self,content):
        ipos = content.find("//")
        if ipos!=-1:
            if content[ipos-1] != ":":
                content = content.replace(content[ipos:],"")
        content = content.replace("\n","")
        if self.FileType != "MENU.txt" and self.FileType != "ROOT.txt":
            content = content.strip()
        else:
            myData = content
            if (myData.strip()==""):
                return ""
            content = content
        return content
        
    def CNDTCToOther(self):
        #self.DTC2dict()
        try:
            with open(self.CNFile,r"r",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Read:
                CNLines = CN_Read.readlines()
                total_files = len(CNLines)
                self.progress["maximum"] = total_files
                self.progress["value"] = 0
                with open(self.OutFile,r"w",encoding = self.encodingDict[self.TargeLanguage], errors='replace') as Targe_Write:
                    for CNLine in CNLines:
                        CNLine = self.Myfilter(CNLine)
                        #空行、注释行
                        if CNLine=="" or CNLine.find("/*")!=-1 or CNLine.find("*/")!=-1:
                            continue
                        if CNLine.find("include")!=-1:
                            Targe_Write.write(CNLine + "\n")
                            continue
                        DTC_List = CNLine.split("\t")
                        Index = DTC_List[0]
                        if contains_chinese(DTC_List[1]) == True:
                            CNValue = re.sub(r"\"$|^\"","",DTC_List[1]).strip()
                            ENValue = self.getTargeValue(CNValue)
                            DTC_List[1] = "\"" + ENValue + "\""
                        PCBU = '\t' + DTC_List[1]
                        MyRe = r"\"$|^\""
                        ENValueAll = ""
                        for i in range(2,len(DTC_List)):
                            CNValue = re.sub(MyRe,"",DTC_List[i]).strip()
                            if CNValue!="" and contains_chinese(CNValue) == True:
                                ENValue = self.getTargeValue(CNValue)
                            else:
                                ENValue = CNValue
                            ENValueAll = ENValueAll +'\t' + "\"" + ENValue + "\""
                        Targe_Write.write(Index + PCBU + ENValueAll + "\n")
                        self.progress["value"] += 1
                        #self.MyGUI.update()
        except Exception as e:
            self.UpdataLog(self.CNFile+" 错误信息："+ str(e) + CNLine)
    def DTC2dict(self):
        CN_dict = {}
        Targe_dict = {}
        self.error = "无法打开文件"
        try:
            with open(self.CNFile,r"r",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Read:
                for CNLine in CN_Read.readlines():
                    self.error = CNLine
                    CNLine = self.Myfilter(CNLine)
                    if CNLine=="" or CNLine.find("include")!=-1:
                        continue
                    CNLineList = CNLine.split("\t")
                    if len(CNLineList)>=3:
                        key = CNLineList[0]
                        value = '\t'.join(CNLineList[2:])
                        # 判断并添加‘*’标记直到没有重复的
                        while key in CN_dict:
                            key += '*'  # 在键的末尾添加‘*’
                        CN_dict[key] = value  # 将最终的唯一键与值插入到字典中
                    else:
                        self.UpdataLog(self.FileType+" 格式错误 "+CNLine)
        except Exception as e:
            self.UpdataLog(self.CNFile+" 错误信息："+self.error+ str(e) + CNLine)
        self.error = "无法打开文件"
        try:
            with open(self.TargeFile,r"r",encoding = self.encodingDict[self.TargeLanguage]) as TargeRead:
                for TargeLine in TargeRead.readlines():
                    self.error = TargeLine
                    TargeLine = self.Myfilter(TargeLine)
                    if TargeLine=="":
                        continue
                    TargeLineList = TargeLine.split("\t")
                    if len(TargeLineList)>=3:
                        Targe_dict[TargeLineList[0].strip()] = TargeLineList[2:].strip()
                    else:
                        self.UpdataLog(self.FileType+" 格式错误 "+TargeLine+ str(e) + CNLine)
        except :
            self.UpdataLog(self.TargeFile+" 错误信息："+self.error + CNLine)

    def CNDTC_HToOther(self):
        if os.path.exists(self.CNFile):
            #self.DTC_H2dict()
            try:
                with open(self.CNFile,r"r",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Read:
                    with open(self.OutFile,r"w",encoding = self.encodingDict[self.TargeLanguage], errors='replace') as Targe_Write:
                        for CNLine in CN_Read.readlines():
                            CNLine = self.Myfilter(CNLine)
                            removeBlank = CNLine.strip()
                            #空行、注释行
                            if removeBlank=="":
                                Targe_Write.write("\n")
                                continue
                            if CNLine.find("include")!=-1:
                                Targe_Write.write(CNLine + "\n")
                                continue
                            DTC_HList = CNLine.split("\t")
                            Index = DTC_HList[0]
                            MyRe = r"\"$|^\""
                            ENValueAll = ""
                            for i in range(1,len(DTC_HList)):
                                CNValue = re.sub(MyRe,"",DTC_HList[i]).strip()
                                if CNValue != "" and contains_chinese(CNValue) == True:
                                    ENValue = self.getTargeValue(CNValue)
                                else:
                                    ENValue = CNValue
                                ENValueAll = ENValueAll +'\t' + "\"" + ENValue + "\""
                            Targe_Write.write(Index + ENValueAll + "\n")
            except Exception as e:
                self.UpdataLog(self.CNFile+" 错误信息："+ str(e))

    def DTC_H2dict(self):
        try:
            Myre = r"\"$|^\""
            CN_dict = {}
            Targe_dict = {}
            self.error = "无法打开文件"
            with open(self.CNFile,r"r",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Read:
                for CNLine in CN_Read.readlines():
                    self.error = CNLine
                    CNLine = self.Myfilter(CNLine)
                    removeBlank = CNLine.strip()
                    #空行、注释行
                    if removeBlank=="":
                        continue
                    if CNLine.find("include")!=-1:
                        continue
                    CNLineList = CNLine.split("\t")
                    if len(CNLineList)==2:
                        if CNLineList[0] in CN_dict:
                            self.UpdataLog(self.FileType+" Index重复 "+CNLine)
                        CN_dict[CNLineList[0]] = re.sub(Myre,"",CNLineList[1]).strip()
                    elif len(CNLineList)==3:
                       if CNLineList[0] in CN_dict:
                           self.UpdataLog(self.FileType+" Index重复 "+CNLine)
                       CN_dict[CNLineList[0].strip()] = CNLineList[2].strip()    
                    else:
                        self.UpdataLog(self.FileType+" 格式错误 "+CNLine)
        except Exception as e:
            self.UpdataLog(self.CNFile+" 错误信息："+self.error+ str(e))
        try:
            self.error = "无法打开文件"
            with open(self.TargeFile,r"r",encoding = self.encodingDict[self.TargeLanguage]) as TargeRead:
                for TargeLine in TargeRead.readlines():
                    self.error = TargeLine
                    TargeLine = self.Myfilter(TargeLine)
                    removeBlank = TargeLine.strip()
                    #空行、注释行
                    if removeBlank=="":
                        continue
                    TargeLineList = TargeLine.split("\t")
                    if len(TargeLineList)==2:
                        if TargeLineList[0] in Targe_dict:
                            self.UpdataLog(self.FileType+" Index重复 "+TargeLine)
                        Targe_dict[TargeLineList[0]] = re.sub(Myre,"",TargeLineList[1]).strip()
                    elif len(TargeLineList)==3:
                        Targe_dict[TargeLineList[0].strip()] = TargeLineList[2].strip()
                    else:
                        self.UpdataLog(self.FileType+" 格式错误 "+TargeLine)
        except:
            self.UpdataLog(self.TargeFile+" 错误信息："+self.error)

    def CNTextToOther(self):
        #self.Text2dict()
        try:
            with open(self.CNFile,r"r",encoding = self.encodingDict["CN-简体"]) as CN_Read:
                with open(self.OutFile,r"w",encoding = self.encodingDict[self.TargeLanguage], errors='replace') as Targe_Write:
                    for CNLine in CN_Read.readlines():
                        CNLine = self.Myfilter(CNLine)
                        removeBlank = CNLine.strip()
                        if CNLine.find("include")!=-1:
                            ipos = CNLine.find("CN")
                            if ipos!=-1:
                                Targe_Write.write(CNLine.replace("CN",self.TargeLanguage[0:2])+"\n")
                            else:
                                Targe_Write.write(CNLine+"\n")
                            continue
                        #空行、注释行
                        if removeBlank=="":
                            Targe_Write.write("\n")
                            continue
                        else:
                            CNLineList = CNLine.split("\t")
                            ENLine = CNLineList[0]+"\t"
                            for i in range(1,len(CNLineList)):
                                MyRe = r"\"$|^\""
                                CNValue = re.sub(MyRe,"",CNLineList[i])
                                CNValue = CNValue.strip()
                                ENValue = "\"\""
                                if CNValue!="" and contains_chinese(CNValue) == True:
                                    ENValue = self.getTargeValue(CNValue)
                                    ENLine = ENLine+ CNLineList[i].replace(CNValue,ENValue)+"\t"
                                else:
                                    ENLine = ENLine+ CNLineList[i]+"\t"
                            ENLine = ENLine[0:-1]
                            Targe_Write.write(ENLine + "\n")
        except Exception as e:
            self.UpdataLog(self.CNFile+" 错误信息："+ str(e))

    def Text2dict(self):
        try:
            self.error = "无法打开文件"
            CN_dict = {}
            Targe_dict = {}
            Myre = r"\"$|^\""
            with open(self.CNFile,r"r",encoding = self.encodingDict["CN-简体"]) as CN_Read:
                for CNLine in CN_Read.readlines():
                    if CNLine.find("include")!=-1:
                        continue
                    self.error = CNLine
                    CNLine = self.Myfilter(CNLine)
                    removeBlank = CNLine.strip()
                    #空行、注释行
                    if removeBlank=="":
                        continue
                    CNLineList = CNLine.split("\t")
                    if CNLineList[0] in CN_dict:
                        self.UpdataLog(self.FileType+" Index重复 "+CNLine)
                    if len(CNLineList)==2:
                        CN_dict[CNLineList[0]] = re.sub(Myre,"",CNLineList[1]).strip()
                    elif len(CNLineList)==3:
                        CN_dict[CNLineList[0]] =re.sub(Myre,"",CNLineList[1]).strip()
                        CN_dict[CNLineList[0]+"_1"] =re.sub(Myre,"",CNLineList[2]).strip()
                    else:
                        self.UpdataLog(self.FileType+" 格式错误 "+CNLine)
        except Exception as e:
            self.UpdataLog(self.CNFile+" 错误信息："+self.error + str(e))
        try:
            self.error = "无法打开文件"
            with open(self.TargeFile,r"r",encoding = self.encodingDict[self.TargeLanguage]) as TargeRead:
                for TargeLine in TargeRead.readlines():
                    self.error = TargeLine
                    TargeLine = self.Myfilter(TargeLine)
                    removeBlank = TargeLine.strip()
                    #空行、注释行
                    if removeBlank=="":
                        continue
                    TargeLineList = TargeLine.split("\t")
                    if TargeLineList[0] in Targe_dict:
                        self.UpdataLog(self.FileType+" Index重复 "+TargeLine)
                    if len(TargeLineList)==2:
                        Targe_dict[TargeLineList[0]] = re.sub(Myre,"",TargeLineList[1]).strip()
                    elif len(TargeLineList)==3:
                        Targe_dict[TargeLineList[0]] = re.sub(Myre,"",TargeLineList[1]).strip()
                        Targe_dict[TargeLineList[0]+"_1"] = re.sub(Myre,"",TargeLineList[2]).strip()
                    else:
                        self.UpdataLog(self.FileType+" 格式错误 "+TargeLine)
        except :
            self.UpdataLog(self.TargeFile+" 错误信息："+self.error)
        self.FileToLocalDict(CN_dict,Targe_dict)

    def CQuickInfoToOther(self):
        if os.path.exists(self.CNFile):
            try:
                with open(self.CNFile,r"r",encoding = self.encodingDict["CN-简体"]) as CN_Read:
                    with open(self.OutFile,r"w",encoding = self.encodingDict[self.TargeLanguage], errors='replace') as Targe_Write:
                        for CNLine in CN_Read.readlines():
                            if CNLine.find("GroupName")!=-1 or CNLine.find("SySName")!=-1:
                                ipos = CNLine.find("=")
                                CNValue = CNLine[ipos+1:].strip()
                                ipos = CNValue.find("\t\t")
                                CNValue = CNValue[0:ipos]
                                ENValue = self.getTargeValue(CNValue)
                                Targe_Write.write(CNLine.replace(CNValue,ENValue))
                            else:
                                Targe_Write.write(CNLine)
            except Exception as e:
                self.UpdataLog(self.CNFile+" 错误信息："+ str(e))

    def Menu2dict(self):
        try:
            self.error = "无法打开文件"
            CN_dict = {}
            with open(self.CNFile,r"r",encoding = self.encodingDict["CN-简体"]) as CN_Read:
                 for CNLine in CN_Read.readlines():
                    self.error = CNLine
                    if CNLine.startswith("\t")==False:
                        continue
                    CNLine = self.Myfilter(CNLine)#去除注释、换行
                    CNLine = CNLine.replace("\t","")
                    ipos = CNLine.find("<")
                    if ipos!=-1:
                        CNLine = CNLine[0:ipos]
                    #直接翻译
                    if CNLine!="":
                        self.getTargeValue(CNLine)
                    #if CNLine not in self.CheckedDataDict:
                    #   if CNLine not in self.ExistDataDict:
                    #       if CNLine not in self.YodaoDataDict:
                    #           self.NOTranDataDict[CNLine] = self.FileType
        except Exception as e:
            self.UpdataLog(self.CNFile+" 错误信息："+self.error+ str(e))

    def CMenuToOther(self):
        #self.Menu2dict()
        try:
            with open(self.CNFile,r"r",encoding = self.encodingDict["CN-简体"]) as CN_Read:
                with open(self.OutFile,r"w",encoding = self.encodingDict[self.TargeLanguage], errors='replace') as Targe_Write:
                    for CNLine in CN_Read.readlines():
                        CNLine = self.Myfilter(CNLine)#去除注释、换行
                        sourceLine = CNLine
                        if CNLine == "":
                            #Targe_Write.write("\n")
                            continue
                        if CNLine.startswith("\t") == False:
                            Targe_Write.write(CNLine+"\n")
                            continue
                        CNLine = self.Myfilter(CNLine)
                        NeedTran = CNLine.replace("\t","")
                        ipos = NeedTran.find("<")
                        if ipos!=-1:
                            NeedTran = NeedTran[0:ipos]
                        Targe = self.getTargeValue(NeedTran)
                        Targe = sourceLine.replace(NeedTran,Targe)
                        Targe_Write.write(Targe+"\n")
        except Exception as e:
            self.UpdataLog(self.CNFile+" 错误信息："+ str(e))
    def Root2dict(self):
        self.error = "无法打开文件"
        CN_dict = {}
        Targe_dict = {}
        try:
            with open(self.CNFile,r"r",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Read:
                 for CNLine in CN_Read.readlines():
                    CNLine = self.Myfilter(CNLine)#去除注释、换行
                    CNLine = CNLine.replace("\t","")
                    self.error = CNLine
                    ipos = CNLine.find("<0x")
                    if ipos !=-1:
                        Id = CNLine[ipos:]
                        content = CNLine[0:ipos]
                        CN_dict[Id] = content
                    elif CNLine!="":
                        header = CNLine[0:2]
                        if header.find("@")!=-1:
                            CNLine = CNLine[2:]
                        self.getTargeValue(CNLine)
                        #if CNLine not in self.CheckedDataDict:
                        #    if CNLine not in self.ExistDataDict:
                        #        if CNLine not in self.YodaoDataDict:
                        #            self.NOTranDataDict[CNLine] = self.FileType
                            
        except Exception as e:
            self.UpdataLog(self.CNFile+" 错误信息："+self.error+ str(e))
        try:
            self.error = "无法打开文件"
            with open(self.TargeFile,r"r",encoding = self.encodingDict[self.TargeLanguage]) as TargeRead:
                 for TargeLine in TargeRead.readlines():
                    self.error = TargeLine
                    TargeLine = self.Myfilter(TargeLine)#去除注释、换行
                    TargeLine = TargeLine.replace("\t","")
                    ipos = TargeLine.find("<0x")
                    if ipos !=-1:
                        Id = TargeLine[ipos:]
                        content = TargeLine[0:ipos]
                        Targe_dict[Id] = content
        except :
            self.UpdataLog(self.TargeFile+" 错误信息："+self.error)
            
    def CRootToOther(self):
        #self.Root2dict()
        try:
            with open(self.CNFile,r"r",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Read:
                with open(self.OutFile,r"w",encoding = self.encodingDict[self.TargeLanguage], errors='replace') as Targe_Write:
                    for CNLine in CN_Read.readlines():
                        CNLine = self.Myfilter(CNLine)#去除注释、换行
                        sourceLine = CNLine
                        if CNLine == "":
                            #Targe_Write.write("\n")
                            continue
                        CNLine = self.Myfilter(CNLine)
                        NeedTran = CNLine.replace("\t","")
                        header = NeedTran[0:2]
                        if header.find("@")!=-1:
                            NeedTran = NeedTran[2:]
                        ipos = NeedTran.find("<")
                        if ipos!=-1:
                            NeedTran = NeedTran[0:ipos]
                        Targe = self.getTargeValue(NeedTran)
                        Targe = sourceLine.replace(NeedTran,Targe)
                        Targe_Write.write(Targe+"\n")
            #with open(self.DataPath+"CN_NEW"+self.FileType,r"w",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_write:

        except Exception as e:
            self.UpdataLog(self.CNFile+" 错误信息："+ str(e))

    def CNDSToOther(self):
        #self.DS2dict()
        with open(self.CNFile,r"r",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Read:
            with open(self.OutFile,r"w",encoding = self.encodingDict[self.TargeLanguage], errors='replace') as Targe_Write:
                for CNLine in CN_Read.readlines():
                    CNLine = self.Myfilter(CNLine)
                    removeBlank = CNLine.strip()
                    #空行、注释行
                    if removeBlank=="":
                        continue
                    if contains_chinese(removeBlank)==False:
                        Targe_Write.write(CNLine + "\n")
                        continue
                    if CNLine.find("include")!=-1:
                        Targe_Write.write(CNLine + "\n")
                        continue
                    CNLineList = removeBlank.split("\t")
                    MyRe = r"\"$|^\""
                    if len(CNLineList)<6:
                        ENLine = ""
                        for i in range(len(CNLineList)):
                            CNValue = re.sub(MyRe,"",CNLineList[i])
                            if contains_chinese(CNValue)==True:
                                if i == 3 and CNValue.find("|")!=-1:
                                    CNValueList = CNValue.split("|")
                                    ENValue = ""
                                    for CNOneValue in CNValueList:
                                        CNOneValue = CNOneValue.strip()
                                        if CNOneValue!="":
                                            CNOneValue = self.getTargeValue(CNOneValue)
                                        ENValue = ENValue + CNOneValue
                                        ENValue = ENValue+"|"
                                    ENValue = ENValue[0:-1]
                                    CNLineList[3] = CNLineList[3].replace(CNValue,ENValue)
                                else:
                                    ENValue = self.getTargeValue(CNValue)
                                    CNLineList[i] = CNLineList[i].replace(CNValue,ENValue)
                        for i in CNLineList:
                            ENLine = ENLine+ i +"\t"
                        ENLine = ENLine[0:-1]
                        Targe_Write.write(ENLine+"\n")
                    else:
                        CNName = re.sub(MyRe,"",CNLineList[1])
                        ENName = ""
                        if CNName!="":
                            ENName = self.getTargeValue(CNName)
                        CNUnit = CNLineList[2]
                        if re.sub(MyRe,"",CNUnit).strip() in self.UNIT_dict:
                            ENUnit = "\"" +self.UNIT_dict[re.sub(MyRe,"",CNUnit).strip()] +"\""
                        elif contains_chinese(re.sub(MyRe,"",CNUnit).strip())==True:
                            ENUnit = self.getTargeValue(re.sub(MyRe,"",CNUnit).strip())
                            ENUnit = "\"" +ENUnit +"\""
                            self.UNIT_dict[CNUnit] = ENUnit
                        else:
                            ENUnit = "\"" +re.sub(MyRe,"",CNUnit).strip() +"\""
                            self.UNIT_dict[ENUnit] = ENUnit
                        CNMore = re.sub(MyRe,"",CNLineList[4])
                        if contains_chinese(CNMore)==True:
                            ENValue = ""
                            CNMores = CNMore.split(";")
                            if len(CNMores)>1:
                                for CNValue in CNMores:
                                    #获取CNMone中最后一个空格的位置
                                    ipos = CNValue.rfind(" ")
                                    if ipos!=-1:
                                        CNValue = CNValue[ipos+1:]
                                        ENValue = self.getTargeValue(CNValue)
                                        CNLineList[4] = CNLineList[4].replace(CNValue,ENValue)
                            else:
                                CNMores = extract_chinese_from_ternary(CNMore)
                                if len(CNMores)>0:
                                    for CNValue in CNMores:
                                        ENValue = self.getTargeValue(CNValue)
                                        CNLineList[4] = CNLineList[4].replace(CNValue,ENValue)
                        try:
                            CNValue = re.sub(MyRe,"",CNLineList[3])
                            CNValueList = CNValue.split("|")
                            ENValue = ""
                            for CNOneValue in CNValueList:
                                CNOneValue = CNOneValue.strip()
                                
                                if CNOneValue.find("%.")!=-1 or CNOneValue.find("%d")!=-1:
                                    ENValue = CNOneValue
                                else:
                                    if CNOneValue!="":
                                        CNOneValue = self.getTargeValue(CNOneValue)
                                    ENValue = ENValue + CNOneValue
                                ENValue = ENValue+"|"
                            ENValue = ENValue[0:-1]
                            ENLine = ""
                            CNLine = ""
                            for i in CNLineList:
                                CNLine = CNLine+ i +"\t"
                            CNLineList[1] = CNLineList[1].replace(CNName,ENName)
                            CNLineList[2] = CNLineList[2].replace(CNUnit,ENUnit)
                            CNLineList[3] = CNLineList[3].replace(CNValue,ENValue)
                            for i in CNLineList:
                                ENLine = ENLine+ i +"\t"
                            ENLine = ENLine[0:-1]
                            CNLine = CNLine[0:-1]
                            Targe_Write.write(ENLine+"\n")
                        except Exception as e:
                            self.UpdataLog("CN_DS、"+CNLine+" 格式有问题"+ str(e))                        

    def DS2dict(self):
        try:
            self.error = "无法打开文件"
            CN_dict = {}
            Targe_dict = {}
            with open(self.CNFile,r"r",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Read:
                for CNLine in CN_Read.readlines():
                    self.error = CNLine
                    CNLine = self.Myfilter(CNLine)
                    removeBlank = CNLine.strip()
                    #空行、注释行
                    if removeBlank=="":
                        continue
                    if CNLine.find("include")!=-1:
                        continue
                    CNLineList = CNLine.split("\t")

                    if CNLineList[0] in CN_dict:
                        self.UpdataLog(self.FileType+" Index重复 "+CNLine)
                    if len(CNLineList)>3:
                        Index = CNLineList[0]
                        value = '\t'.join(CNLineList[1:])
                        CN_dict[Index] = value
                    else:
                        self.UpdataLog(self.FileType+" 格式错误 ",CNLine)
        except Exception as e:
            self.UpdataLog(self.CNFile+" 错误信息："+self.error+ str(e))
        try:
            self.error = "无法打开文件"
            with open(self.TargeFile,r"r",encoding = self.encodingDict[self.TargeLanguage], errors='replace') as TargeRead:
                for TargeLine in TargeRead.readlines():
                    self.error = TargeLine
                    TargeLine = self.Myfilter(TargeLine)
                    removeBlank = TargeLine.strip()
                    #空行、注释行
                    if removeBlank=="":
                        continue
                    TargeLineList = TargeLine.split("\t")
                    if TargeLineList[0] in Targe_dict:
                        self.UpdataLog(self.FileType+" Index重复 "+TargeLine)
                    if len(TargeLineList)>3:
                        Index = TargeLineList[0]
                        value = '\t'.join(TargeLineList[1:])
                        Targe_dict[Index] = value
                    else:
                        self.UpdataLog(self.FileType+" 格式错误 "+TargeLine)
        except:
            self.UpdataLog(self.TargeFile+" 错误信息："+self.error)
        self.FileToLocalDict(CN_dict,Targe_dict)

    def read_cell(self,sheetName,col):
        CN_sheet = self.CN_work.sheet_by_name(sheetName)
        self.New_sheet = self.New_work.get_sheet(sheetName)
        CN_rows = CN_sheet.nrows
        for row in range(1,CN_rows):
            my_cell_value = CN_sheet.cell_value(row,col)
            strid = str(CN_sheet.cell_value(row,0)).replace(".0","").strip()
            if len(my_cell_value)==0:
                continue
            elif col ==3 and (sheetName=="ReadCds" or sheetName=="FreezeFrame"):
                self.write_xls(strid,row,col,my_cell_value,sheetName,True)
            else:
                self.write_xls(strid,row,col,my_cell_value,sheetName)

    def write_xls(self,strid,row,col,value,SheetName,Unit = False):
        New_word = ""
        if Unit:#单位翻译
            if value in self.UNIT_dict:
                New_word = self.UNIT_dict[value]
            else:
                self.UNIT_dict[value] = value
                self.UpdataLog(SheetName+" "+strid+" 新的单位出现："+value)
                return 0
        elif col==5 and (SheetName=="ReadCds" or SheetName=="FreezeFrame"):#数据流格式翻译
            valueList = value.split("|")
            NewValueList = []
            if self.TargeXlsExsit:
                if len(valueList)!=len(self.ENExcelDict[SheetName]["Value"][strid]):
                    self.UpdataLog("中英文库的值不一样："+self.CNPath+"/t"+SheetName+"/t"+strid)
                    for oneValue in valueList:
                        XLSOneValue = oneValue
                        if oneValue.find("%")!=-1:
                            NewValueList.append(oneValue+"|")
                        else:
                            if XLSOneValue!="":
                                XLSOneValue = self.getTargeValue(XLSOneValue)
                            NewValueList.append(XLSOneValue+"|")
                else:
                    i = 0
                    for oneValue in valueList:
                        XLSOneValue = oneValue
                        myWord = ""
                        if oneValue.find("%")!=-1:
                            myWord = oneValue
                        else:
                            if oneValue in self.CheckedDataDict:
                                myWord = self.CheckedDataDict[XLSOneValue]
                            if myWord=="" and (oneValue in self.ExistDataDict):
                                myWord = self.ExistDataDict[XLSOneValue]
                            if myWord=="":
                                myWord = self.ENExcelDict[SheetName]["Value"][strid][strid+"_"+str(i)]
                                self.AddValue2Dict(XLSOneValue,myWord)
                            if myWord=="" and (oneValue in self.YodaoDataDict):
                                myWord = self.YodaoDataDict[XLSOneValue]
                            if myWord=="":
                                myWord = self.youdaoTranslate(XLSOneValue)
                        NewValueList.append(myWord+"|")
                        i = i+1
            else:
                for oneValue in valueList:
                    XLSOneValue = oneValue
                    if oneValue.find("%")!=-1:
                        NewValueList.append(oneValue+"|")
                    else:
                        if XLSOneValue!="":
                            XLSOneValue = self.getTargeValue(XLSOneValue)
                        NewValueList.append(XLSOneValue+"|")
            if len(NewValueList)!=0:
                NewValueList[-1] = NewValueList[-1].replace("|","")
            for oneValue in NewValueList:
                New_word = New_word+oneValue
        elif value.find("%")!=-1 and SheetName=="EcuInfo":#信息版本，没有文本
            return 0
        else:#正常串翻译
            XLSOneValue = value
            if XLSOneValue in self.CheckedDataDict:
                New_word = self.CheckedDataDict[XLSOneValue]
            if New_word == "" and (XLSOneValue in self.ExistDataDict):
                New_word = self.ExistDataDict[XLSOneValue]
            if New_word == "":
                if self.TargeXlsExsit:
                    if SheetName=="EcuInfo" or SheetName=="Text" or SheetName=="Stat":
                        if strid in self.ENExcelDict[SheetName]:
                            New_word = self.ENExcelDict[SheetName][strid]
                            if New_word!="":
                                self.AddValue2Dict(XLSOneValue,New_word)
                        if New_word=="" and (XLSOneValue in self.YodaoDataDict):
                            New_word = self.YodaoDataDict[XLSOneValue]
                        if New_word=="":
                            New_word = self.youdaoTranslate(XLSOneValue,self.CNPath)
                    elif SheetName=="ReadCds" or SheetName=="FreezeFrame":
                        if strid in self.ENExcelDict[SheetName]["Name"]:
                            New_word = self.ENExcelDict[SheetName]["Name"][strid]
                            if New_word!="":
                                self.AddValue2Dict(XLSOneValue,New_word)
                        if New_word=="" and (XLSOneValue in self.YodaoDataDict):
                            New_word = self.YodaoDataDict[XLSOneValue]
                        if New_word=="":
                            New_word = self.youdaoTranslate(XLSOneValue,self.CNPath)
                    elif SheetName=="Dtc":
                        if strid in self.ENExcelDict[SheetName]["Description"]:
                            if col==2:
                                New_word = self.ENExcelDict[SheetName]["Description"][strid]
                            else:
                                New_word = self.ENExcelDict[SheetName]["Help"][strid]
                            if len(New_word)!=0:
                                self.AddValue2Dict(XLSOneValue,New_word)
                            else:
                                if XLSOneValue in self.YodaoDataDict:
                                    New_word = self.YodaoDataDict[XLSOneValue]
                                if New_word == "":
                                    New_word = self.youdaoTranslate(XLSOneValue,self.CNPath)
                        elif XLSOneValue in self.YodaoDataDict:
                            New_word = self.YodaoDataDict[XLSOneValue]
                        if New_word=="":
                            New_word = self.youdaoTranslate(XLSOneValue,self.CNPath)
                
                if New_word=="" and XLSOneValue!="":
                    New_word = self.getTargeValue(XLSOneValue)
                    
        self.New_sheet.write(row,col,New_word)#重写

    def UpdataLog(self,CurrentInfo):
        self.logInfo.append(CurrentInfo + "\n")
        self.LogBox.insert(END,CurrentInfo + "\n")
        self.LogBox.see(END)
        self.LogBox.update()
        self.MyGUI.update()

    def FileToLocalDict(self,CN_dict,Targe_dict):
        MyRe = r"\"$|^\""#去除字符串两端的冒号
        for key,value in CN_dict.items():
            if self.FileType=="DS.txt":
                CNValueList = value.split("\t")
                CNName = re.sub(MyRe,"",CNValueList[0]).strip()
                CNUnit = re.sub(MyRe,"",CNValueList[1]).strip()
                CNSubValue = re.sub(MyRe,"",CNValueList[2])
                if key in Targe_dict:#有对应的翻译
                    ENValueList = Targe_dict[key].split("\t")
                    ENName = re.sub(MyRe,"",ENValueList[0]).strip()
                    self.AddValue2Dict(CNName,ENName)
                    ENUnit = re.sub(MyRe,"",ENValueList[1]).strip()
                    if CNUnit not in self.UNIT_dict:
                        self.UpdataLog(self.FileType+" 新单位出现："+key)
                        self.UNIT_dict[CNUnit] = ENUnit
                    ENSubValue = re.sub(MyRe,"",ENValueList[2])
                    CNSubValueList = CNSubValue.split("|")
                    ENSubValueList = ENSubValue.split("|")
                    if len(CNSubValueList)!=len(ENSubValueList):
                        self.UpdataLog(self.FileType+"Index {strkey} 中英文库的值不一样".format(strkey = key))
                        if CNSubValueList[i].strip() not in self.ExistDataDict:
                            self.youdaoTranslate(CNSubValueList[i].strip())
                    else:
                        for i in range(len(CNSubValueList)):
                            if CNSubValueList[i].find("%")!=-1:#特殊字符不翻译
                                pass
                            else:
                                if CNSubValueList[i].strip() not in self.ExistDataDict:
                                    self.AddValue2Dict(CNSubValueList[i].strip(),ENSubValueList[i].strip())
                else:#无对应翻译
                    if CNName in self.ExistDataDict:
                        pass
                    else:
                        self.youdaoTranslate(CNName)
                            
                    if CNUnit not in self.UNIT_dict:
                        self.UpdataLog(self.FileType+" 新单位出现："+key)
                        self.UNIT_dict[CNUnit] = CNUnit
                    CNSubValueList = CNSubValue.split("|")
                    for CNOneSubValue in CNSubValueList:
                        if CNOneSubValue.find("%")!=-1:#特殊字符不翻译
                            continue
                        elif CNOneSubValue in self.ExistDataDict:
                            continue
                        else:
                            self.youdaoTranslate(CNOneSubValue)
            elif self.FileType=="TEXT.txt" or self.FileType=="DTC.txt" or self.FileType=="ROOT.txt" or self.FileType=="DTC_H.txt":
                CNValue = re.sub(MyRe,"",value).strip()
                ENValue = ""
                if key in Targe_dict:
                    ENValue = re.sub(MyRe,"",Targe_dict[key]).strip()
                    self.AddValue2Dict(CNValue,ENValue)
                else:
                    self.youdaoTranslate(CNValue)

    def checkChar(self,text):
        if len(text)==0:
            return 2
        for i in range(0,len(text)):
            if 0<ord (text[i]) and ord (text[i])>127:#英语字符范围0-127
                #print(text[i])
                return 0
        return 1

    def youdaoTranslate(self,SorseText,xls=""):
        ENText = ""
        if SorseText in self.CheckedDataDict:
            ENText = self.CheckedDataDict[SorseText]
            if ENText!="":
                return ENText
        if SorseText in self.ExistDataDict:
            ENText = self.ExistDataDict[SorseText]
            if ENText!="":
                return ENText
        if SorseText in self.YodaoDataDict:
            ENText = self.YodaoDataDict[SorseText]
            if ENText!="":
                #self.UpdataLog("已有翻译："+SorseText+"->"+ENText)
                return ENText
        
        if self.checkChar(SorseText)==1 or len(SorseText)==0:
            #self.UpdataLog("实时翻译："+SorseText+"->"+SorseText)
            self.YodaoDataDict[SorseText]=SorseText
            return SorseText
        ENText = self.NoTranFlah
        if self.Toyoudao:
            #ENText = self.Mytran(SorseText)
            #if ENText!=SorseText:
            #    self.YodaoDataDict[SorseText]=ENText
            #else:
            #    self.NOTranDataDict[SorseText]=self.FileType
            ENText = gooletranslate(SorseText,target_language = self.TargeLanguage)
            #self.UpdataLog("实时翻译："+SorseText+"->"+ENText)
            self.MyGUI.update()
            self.YodaoDataDict[SorseText]=ENText
        else:
            self.NOTranDataDict[SorseText]= self.FileType+" " +xls
        return ENText

    def Mytran(self,cn):
        #print(pyautogui.position())
        #time.sleep(100)
        for i in range(0,1):
            i=0
            en = cn
            MyClipboard.copy(cn)
            pyautogui.moveTo(171,1061)
            pyautogui.click()
            pyautogui.moveTo(174,659)
            pyautogui.click()
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.press('backspace')
            pyautogui.hotkey('ctrl', 'v')
            time.sleep(0.5+i)#翻译延迟
            pyautogui.moveTo(442,1014)
            pyautogui.click()
            time.sleep(0.2+i)#翻译延迟
            en = MyClipboard.paste()
            pyautogui.moveTo(171,1061)
            pyautogui.click()
            if en!=cn:
                return en
            i= i + 1
        return cn

    def getTargeValue(self,CN):
        EN = ""
        if CN.find("include")!=-1:
            return CN
        if CN in self.CheckedDataDict:
            EN = self.CheckedDataDict[CN]
            if len(EN)==0:
                if CN in self.ExistDataDict:
                    EN = self.ExistDataDict[CN]
                    if len(EN)==0:
                        if CN in self.YodaoDataDict:
                            EN = self.YodaoDataDict[CN]
        if len(EN)==0:
            EN = self.youdaoTranslate(CN)
        if len(EN)==0:
            if contains_chinese(CN)==False:
                EN = CN
            EN = self.NoTranFlah
        return EN

    def mylog(self):
        current_time = self.get_current_time().strip().replace(" ","_").replace("-","_").replace(":","_")+".txt"
        folder = os.path.exists("Log")
        if not folder:
            os.makedirs("Log")
        with open("Log//"+current_time,'w',encoding = "utf-8")as lgwrite:
            for line in self.logInfo:
                lgwrite.write(line)
        self.logInfo.clear()

    def CheckReadDtc(self):
        CN_sheet = self.CN_work.sheet_by_name("ReadDtc")
        CN_rows = CN_sheet.nrows
        for row in range(0,CN_rows):
            list5 = str(CN_sheet.cell_value(row,5)).replace(".0","").strip()
            list6 = str(CN_sheet.cell_value(row,6)).replace(".0","").strip()
            list7 = str(CN_sheet.cell_value(row,7)).replace(".0","").strip()
            list8 = str(CN_sheet.cell_value(row,8)).replace(".0","").strip()
            if list5=="3" and list6=="4" and list8!="4" and list7.find("[2]")!=-1:
                self.UpdataLog("存在问题："+self.CNPath)
                break

    def get_current_time(self):
        current_time = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
        return current_time

    def CheckGarbledCode(self):
        self.strOutPath="36h.txt"
        self.LangueSelect.update()
        self.TargeLanguage = self.LangueSelect.get()
        if self.TargeLanguage=="EN-英语":
            self.creatFile()
            self.xls2dict()
            Warn = "存在异常："
            i=1
            for k,v in self.CheckedDataDict.items():
                if i:
                    i=0
                    continue
                if self.checkChar(v)!=1 and len(k)!=0:
                    self.UpdataLog(Warn+k+"->"+v)
                    
            i=1
            for k,v in self.ExistDataDict.items():
                if i:
                    i=0
                    continue
                if self.checkChar(v)!=1 and len(k)!=0:
                    self.UpdataLog(Warn+k+"->"+v)
            i=1
            for k,v in self.YodaoDataDict.items():
                if i:
                    i=0
                    continue
                if self.checkChar(v)!=1 and len(k)!=0:
                    self.UpdataLog(Warn+k+"->"+v)
            i=1
            for k,v in self.UNIT_dict.items():
                if i:
                    i=0
                    continue
                if self.checkChar(v)==0:
                    self.UpdataLog("单位乱码："+k+"->"+v)
        else:
            self.UpdataLog("未支持该语言检查")
    
    def CheckFormatStart(self):
        if self.strVehicel=="":
            return 0
        self.ObjSelect.update()
        if self.ObjSelect.get()=="所有.txt":
            FileTypeList = ["DS.txt","DTC_H.txt","DTC.txt","TEXT.txt"]
        else:
            FileTypeList = [self.ObjSelect.get()]
        
        for FileType in FileTypeList:
            self.FileType = FileType
            self.CheckFun()
        self.UpdataLog("********************************格式调整完成，建议用Beyond Compare对比检查一下********************************")
        self.mylog()

    def CheckFun(self):
        self.UpdataLog("*******************************************************开始检查："+self.FileType)
        self.CNFile = self.DataPath+"CN_"+self.FileType
        self.CNFileNew = self.DataPath+"CN_"+"NEW_"+self.FileType
        self.CheckFormat()
        self.UpdataLog("*******************************************************检查完成："+self.FileType+"\n")

    def CheckFormat(self):
        try:
            with open(self.CNFile,r"r",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Read:
                CN_Lines = CN_Read.readlines()
                total_files = len(CN_Lines)
                self.progress["maximum"] = total_files
                self.progress["value"] = 0
                with open(self.CNFileNew,r"w",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Write:
                    for line in CN_Lines:
                        #line = self.Myfilter(line)
                        line = line.strip()
                        if line.find("include")!=-1 or line=="" or line=="//":
                            CN_Write.write(line + "\n")
                            continue
                        CheckLine = process_string(line)
                        if has_invalid_characters(CheckLine, self.encodingDict["CN-简体"]):
                            self.UpdataLog(CheckLine+" 存在乱码")
                        CheckLine = self.replace_chinese_symbol(CheckLine)
                        CN_Write.write(CheckLine + "\n")
                        self.progress["value"] += 1
        except FileNotFoundError:
            self.UpdataLog(f"文件未找到: {self.CNFile}，跳过此文件。\n")  # 记录日志，表示文件未找到
            return  # 跳过后退出函数
    
    def GetTransInfo(self):
        if self.CheckAll.get() == 1:
            self.GetTransInfoAll()
        else:
            if self.strVehicel=="":
                return 0
            self.CNFileNew = self.DataPath+"CN_NeedTrans.txt"
            if os.path.exists(self.CNFileNew):
                os.remove(self.CNFileNew)
            self.ObjSelect.update()
            if self.ObjSelect.get()=="所有.txt":
                FileTypeList = ["ROOT.txt","MENU.txt","DS.txt","DTC_H.txt","DTC.txt","TEXT.txt","QuickInfo.txt"]
            else:
                FileTypeList = [self.ObjSelect.get()]
            empty_set = set()
            for FileType in FileTypeList:
                self.FileType = FileType
                self.GetFun(empty_set)
            with open(self.CNFileNew,r"w",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Write:
                for k in empty_set:
                    CN_Write.write(k + "\n")
            self.UpdataLog("********************************全部提取完成********************************")
            self.mylog()

    def GetTransInfoAll(self):
        #获取文件夹内的所有文件,如果还有子文件夹，递归
        file_list = []
        empty_set = set()
        try:#没有选择文件夹
            for root, dirs, files in os.walk(self.DataPath):
                for file in files:#.txt文件
                    if (file.endswith(".txt") and file.find("_NeedTrans.txt") == -1) or file.endswith(".ini"):
                        file_list.append(os.path.join(root, file))
        except:
            self.UpdataLog("********************************未选择文件夹********************************")
            return 0
        #获取所有需要翻译的文件
        total_files = len(file_list)
        self.progress["maximum"] = total_files
        self.progress["value"] = 0
        self.UpdataLog("*******************************************************开始提取")
        for file_name in file_list:
            if file_name.endswith(".ini"):
                try:
                    with open(file_name,r"r",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Read:
                        CN_Lines = CN_Read.readlines()
                        for line in CN_Lines:
                            CN_Values = line.split("\t")
                            if contains_chinese(CN_Values[1]):
                                empty_set.add(CN_Values[1].strip())
                except:
                    continue  # 跳过后继续下一个文件
            else:
                try:#打开文件
                    with open(file_name,r"r",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Read:
                        CN_Lines = CN_Read.readlines()
                        for line in CN_Lines:
                            if contains_chinese(line):
                                #获取line中的中文字符串
                                chinese_substrings = extract_chinese_from_ternary(line)
                                for substring in chinese_substrings:
                                    empty_set.add(substring.strip())
                except:
                    continue  # 跳过后继续下一个文件
            self.progress["value"] += 1
            self.MyGUI.update()
        with open(self.DataPath+"CN_NeedTrans.txt",r"w",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Write:
            for k in empty_set:
                CN_Write.write(k + "\n")
        self.UpdataLog("*******************************************************提取完成")

    def GetFun(self, empty_set):
        self.UpdataLog("*******************************************************开始提取："+self.FileType)
        self.CNFile = self.DataPath+"CN_"+self.FileType
        self.Get(empty_set)
        self.UpdataLog("*******************************************************提取完成："+self.FileType+"\n")

    def Get(self, empty_set):
        try:
            with open(self.CNFile,r"r",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Read:
                CN_Lines = CN_Read.readlines()
                total_files = len(CN_Lines)
                self.progress["maximum"] = total_files
                self.progress["value"] = 0
                for line in CN_Lines:
                    line = self.Myfilter(line)
                    line = line.strip()
                    if line.find("include")!=-1 or line=="" or line.find("#define")!=-1:
                        continue
                    header = line[0:2]
                    if header.find("//")!=-1:
                        continue
                    CNLists = line.split("\t")
                    MyRe = r"\"$|^\""
                    for CN in CNLists:
                        index_position = CNLists.index(CN)
                        if self.FileType=="MENU.txt":
                            CNValue = CN
                            ipos = CNValue.find("<")
                            if ipos!=-1:
                                CNValue = CNValue[0:ipos]
                            empty_set.add(CNValue.strip())
                        elif self.FileType=="ROOT.txt":
                            CNValue = CN
                            header = CNValue[0:2]
                            if header.find("@")!=-1:
                                CNValue = CNValue[2:]
                            ipos = CNValue.find("<")
                            if ipos!=-1:
                                CNValue = CNValue[0:ipos]
                            empty_set.add(CNValue.strip())
                        elif self.FileType=="DS.txt" and index_position>0:
                            CNValue = re.sub(MyRe,"",CN.strip())
                            if CNValue.find("|")!=-1:
                                CNValueN = CNValue.split("|")
                                for k in CNValueN:
                                    empty_set.add(k.strip())
                            if index_position==2 or index_position==1:
                                empty_set.add(CNValue.strip())
                            if index_position==4:
                                if contains_chinese(CNValue):
                                    CNValues = CNValue.split(";")
                                    if len(CNValues)>1:
                                        for k in CNValues:
                                            ipos = k.rfind(" ")
                                            if ipos!=-1:
                                                empty_set.add(k[ipos+1:].strip())
                                    else:
                                        CNValues = extract_chinese_from_ternary(CNValue)
                                        if len(CNValues)>0:
                                            for k in CNValues:
                                                empty_set.add(k.strip())
                        elif self.FileType=="QuickInfo.txt":
                            if CN.find("GroupName")!= -1 or CN.find("SySName")!=-1:
                                ipos = CN.find("=")
                                CNValue = CN[ipos+1:].strip()
                                empty_set.add(CNValue.strip())
                        else:
                            CNValue = re.sub(MyRe,"",CN.strip())
                            if contains_chinese(CNValue):
                                empty_set.add(CNValue.strip())
                    self.progress["value"] += 1
                    self.MyGUI.update()
        except FileNotFoundError:
            self.UpdataLog(f"文件未找到: {self.CNFile}，跳过此文件。\n")  # 记录日志，表示文件未找到
            return  # 跳过后退出函数
    
    def GetIDM(self):
        empty_set = set()
        with open(self.DataPath+"CN_ROOT.txt",r"r",encoding = self.encodingDict["CN-简体"], errors='replace') as CN_Read:
            cn_lines = CN_Read.readlines()
            for line in cn_lines:
                hearder = line[0:2]
                if hearder.find("//")!=-1:
                    continue
                line = line.strip()
                ipos = line.find("<")
                if ipos!=-1:
                    IDM = line[ipos+3:ipos+5]
                    empty_set.add(IDM)
        with open(self.DataPath+"IDM.txt",r"w",encoding = self.encodingDict["CN-简体"], errors='replace') as IDM_Write:
            for IDM in empty_set:
                IDM_Write.write(IDM + "\n")
        self.UpdataLog("********************************IDM提取完成********************************")
        self.mylog()

    def load_chinese_to_english_mapping(self):
        """
        加载中文到英文字符的映射表，如果文件不存在则创建并写入映射表。
        """
        mapping_file_path = r"C:\Project Trans\chinese_to_english.json"

        # 检查文件是否存在，如果不存在则创建并写入映射表
        if not os.path.exists(mapping_file_path):
            # 定义中文字符到英文字符的映射
            chinese_to_english = {
                '，': ',',
                '。': '.',
                '：': ':',
                '；': ';',
                '？': '?',
                '！': '!',
                '（': '(',
                '）': ')',
                '【': '[',
                '】': ']',
                '《': '<',
                '》': '>',
                '——': '-',
                '～': '~',
                '、': ',',
                '‘': "'",
                '’': "'",
                '％': "%",
                '‖': "|",
                # 希腊字母
                'α': 'Alpha',
                'β': 'Beta',
                'γ': 'Gamma',
                'δ': 'Delta',
                'ε': 'Epsilon',
                'ζ': 'Zeta',
                'η': 'Eta',
                'θ': 'Theta',
                'ι': 'Iota',
                'κ': 'Kappa',
                'λ': 'Lambda',
                'μ': 'Mu',
                'ν': 'Nu',
                'ξ': 'Xi',
                'ο': 'Omicron',
                'π': 'Pi',
                'ρ': 'Rho',
                'σ': 'Sigma',
                'τ': 'Tau',
                'υ': 'Upsilon',
                'φ': 'Phi',
                'χ': 'Chi',
                'ψ': 'Psi',
                'ω': 'Omega',
                # 单位符号
                '°C': 'degC',
                '°F': 'degF',
                '℃': 'degC',
                '℉': 'degF',
                'μ': 'u',
                'Ω': 'Ohm'
            }
            # 确保目录存在
            os.makedirs(os.path.dirname(mapping_file_path), exist_ok=True)
            # 将映射表写入文件
            with open(mapping_file_path, 'w', encoding='utf-8') as f:
                json.dump(chinese_to_english, f, ensure_ascii=False, indent=4)
            return chinese_to_english
        else:
            # 从文件中读取映射表
            with open(mapping_file_path, 'r', encoding='utf-8') as f:
                return json.load(f)

    def replace_chinese_symbol(self, text):
            """
            将字符串中的中文字符替换为对应的英文字符。

            :param text: 输入的字符串
            :return: 替换后的字符串
            """
            # 直接使用内存中的映射表
            for chinese_char, english_char in self.chinese_to_english.items():
                text = text.replace(chinese_char, english_char)

            return text

def replace_exact_chinese(line, CN_Value, EN_Value):
    # 使用字符串操作确保只替换完全匹配的CN_Value
    start = 0
    while True:
        # 查找CN_Value在line中的位置
        start = line.find(CN_Value, start)
        if start == -1:
            break
        
        # 检查CN_Value是否前后都是非中文字符
        if (start == 0 or not '\u4e00' <= line[start - 1] <= '\u9fa5') and \
           (start + len(CN_Value) == len(line) or not '\u4e00' <= line[start + len(CN_Value)] <= '\u9fa5'):
            # 检查前后是否是空格或符号
            prefix = ' ' if (start != 0 and line[start - 1] not in ' \t\n\r\f\v.,!?;:"|' ) else ''
            suffix = ' ' if (start + len(CN_Value) != len(line) and line[start + len(CN_Value)] not in ' \t\n\r\f\v.,!?;:"|' ) else ''
            
            # 进行替换
            line = line[:start] + prefix + EN_Value + suffix + line[start + len(CN_Value):]
            # 更新start位置以避免无限循环
            start += len(prefix) + len(EN_Value) + len(suffix)
        else:
            # 更新start位置继续查找
            start += len(CN_Value)
    
    return line

def extract_chinese_from_ternary(ternary_expression):

    # 正则表达式匹配中文字符
    pattern = r'[\u4e00-\u9fa5]+'
    chinese_substrings = re.findall(pattern, ternary_expression)
    return chinese_substrings

def contains_chinese(text):
    # 定义正则表达式，匹配中文字符
    pattern = re.compile(r'[\u4e00-\u9fa5]')
    return bool(pattern.search(text))

def process_string(input_string):
    # 定义一个函数用来处理引号中的内容
    def clean_quoted_content(match):
        # 获取引号中的内容
        content = match.group(1)
        # 去除首尾空格，并将连续的空格替换为一个空格
        cleaned_content = re.sub(r'\s+', ' ', content.strip())
        return f'"{cleaned_content}"'  # 重新加上引号
    
    # 将连续的空格（两个或更多）替换为一个制表符
    processed_string = re.sub(r' {2,}', '\t', input_string)

    # 将连续的制表符替换为一个制表符
    processed_string = re.sub(r'\t+', '\t', processed_string)

    # 使用正则表达式匹配引号中的内容并进行处理
    processed_string = re.sub(r'"([^"]*)"', clean_quoted_content, processed_string)
    
    # 更新引号内部的内容，将符号","替换为制表符
    processed_string = re.sub(r'"\s*,\s*"', '"\t"', processed_string)

    # 更新引号内部的内容，将符号" "替换为制表符
    processed_string = re.sub(r'"\s* \s*"', '"\t"', processed_string)

    # 更新第一个引号前的空白字符，将符号" "替换为制表符
    processed_string = re.sub(r'\s+"', '\t"', processed_string)

    # 将处理后的字符串按制表符切割
    parts = processed_string.split('\t')
    
    # 如果parts中的第一个元素中存在引号，在第一个引号前添加一个制表符，并生成新的列表
    if '"' in parts[0]:
        ipos = parts[0].find('"')
        parts[0] = parts[0][:ipos] + '\t' + parts[0][ipos:]
        processed_string = '\t'.join(parts)
        parts = processed_string.split('\t')

    # 去除每部分的前后空白字符
    cleaned_parts = [part.strip() for part in parts]

    # 去除首尾的逗号
    cleaned_parts = [part.strip(',') for part in parts]

    # 如果内容长度大于2且有连续的引号，将连续引号替换成一个引号
    # 检查替换后的引号数量是否为单数，如果是单数则不进行替换
    def remove_consecutive_quotes(part):
        if len(part) > 2:
            # 替换连续引号为一个引号
            new_part = re.sub(r'(["\'])(\1)+', r'\1', part)
            # 检查引号数量是否为单数
            if new_part.count('"') % 2 == 0 and new_part.count("'") % 2 == 0:
                return new_part
        return part

    cleaned_parts = [remove_consecutive_quotes(part) for part in cleaned_parts]

    # 重新用制表符连接这些部分
    return '\t'.join(cleaned_parts)

def has_invalid_characters(input_string, encoding='utf-8'):
    try:
        # 尝试将字符串编码为指定的编码格式
        input_string.encode(encoding)
    except UnicodeEncodeError:
        # 如果抛出 UnicodeEncodeError，表示存在不支持的字符
        return True
    return False

def find_char_positions(s, char):
    positions = []
    for index, c in enumerate(s):
        if c == char:
            positions.append(index)
    return positions

def remove_consecutive_spaces_and_tabs(input_string):
    # 使用正则表达式替换连续的空格或制表符
    # ' + '表示一个或多个空格，'\t+'表示一个或多个制表符
    result = re.sub(r'[ \t]+', ' ', input_string)  # 将多个空格或制表符替换为一个空格
    return result.strip()  # 去掉字符串两端的空白字符

if __name__=='__main__':
    translator = myTranslator()  # 实例化 myTranslator 类
    translator.creatUI()  # 创建并显示用户界面
    translator.MyGUI.mainloop()  # 进入消息循环
